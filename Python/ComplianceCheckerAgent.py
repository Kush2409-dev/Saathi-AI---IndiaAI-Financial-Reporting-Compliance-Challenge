"""
NFRA Audit Compliance ‚Äî LangGraph Orchestration Engine
========================================================
Multi-step agentic reasoning over extracted annual report sections
to evaluate compliance against accounting standards and audit rules.

Folder structure expected:
  Extracted_Reports/
    CompanyA_sections/
      Balance Sheet.md
      Statement of Profit and Loss.md
      Independent Auditor's Report.md
      Notes to Financial Statements.md
      ...
    CompanyB_sections/
      ...

Usage:
  python audit_orchestration.py \
      --reports  "Extracted_Reports" \
      --rules    "Accounting Standard - Rules.xlsx" \
      --output   "Compliance Matrix.xlsx"
"""

import os
import re
import json
import shutil
import argparse
from typing import TypedDict, List, Dict, Optional, Annotated
from pathlib import Path
from collections import OrderedDict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv
from langgraph.graph import StateGraph, END
from langchain_openai import AzureChatOpenAI
from langchain_core.messages import SystemMessage

load_dotenv()  # searches current dir, then parent dirs automatically

# Also try explicit paths common in this project structure
for candidate in [Path(".env"), Path("Python/.env"), Path(__file__).resolve().parent / ".env"]:
    if candidate.exists():
        load_dotenv(dotenv_path=candidate, override=True)
        break

# ‚îÄ‚îÄ Validate env vars loaded ‚îÄ‚îÄ
_key = os.getenv("AZURE_OPENAI_API_KEY")
_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
if not _key or not _endpoint:
    raise ValueError(
        f"Azure OpenAI credentials not found.\n"
        f"  AZURE_OPENAI_API_KEY set:  {bool(_key)}\n"
        f"  AZURE_OPENAI_ENDPOINT set: {bool(_endpoint)}\n"
        f"  Looked for .env in: {Path('.env').resolve()}, {Path('Python/.env').resolve()}\n"
        f"  Tip: Place .env in the directory you run the script from."
    )

LLM = AzureChatOpenAI(
    azure_endpoint=_endpoint,
    api_key=_key,
    api_version=os.getenv("AZURE_OPENAI_API_VERSION", "2024-12-01-preview"),
    azure_deployment=os.getenv("AZURE_OPENAI_DEPLOYMENT"),
    temperature=0,
)


# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê
# 2. STATE
# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê

class AuditState(TypedDict):
    # ‚îÄ‚îÄ Identity ‚îÄ‚îÄ
    company_name: str
    rule_id: str
    rule_details: str
    standard_ref: str
    audit_requirement: str

    # ‚îÄ‚îÄ Questions parsed from rules Excel ‚îÄ‚îÄ
    # Each: {"question": str, "target_sections": [str], "financial_check": dict|None}
    parsed_questions: List[Dict]
    current_question_index: int

    # ‚îÄ‚îÄ Report content (loaded from section .md files) ‚îÄ‚îÄ
    available_sections: Dict[str, str]   # {"Balance Sheet": "md content", ...}
    section_filenames: Dict[str, str]    # {"Balance Sheet": "Balance Sheet.md", ...}

    # ‚îÄ‚îÄ Accumulated results ‚îÄ‚îÄ
    question_responses: List[Dict]

    # ‚îÄ‚îÄ Failure logic ‚îÄ‚îÄ
    failure_trigger_text: str
    failure_trigger_met: bool

    # ‚îÄ‚îÄ Final output (‚Üí Compliance Matrix columns) ‚îÄ‚îÄ
    compliance_status: Optional[str]
    summary_finding: Optional[str]
    auditor_oversight: Optional[str]
    reasoning_path: Optional[str]
    evidence_snippet: Optional[str]
    page_ref: Optional[str]
    confidence_score: Optional[int]


# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê
# 3. REPORT LOADER ‚Äî reads section .md files from a company folder
# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê

def load_company_sections(folder_path: Path) -> Dict[str, str]:
    """
    Load all .md files from a company's extracted sections folder.
    Returns: {"Section Name": "markdown content", ...}
    """
    sections = {}
    for md_file in sorted(folder_path.glob("*.md")):
        section_name = md_file.stem  # filename without .md
        content = md_file.read_text(encoding="utf-8").strip()
        if content:
            sections[section_name] = content
    return sections


def discover_companies(reports_dir: Path) -> List[Dict]:
    """
    Discover all company section folders under Extracted_Reports/.
    Skips COMPLETED folder. Any subfolder containing .md files is a company.
    """
    companies = []
    for item in sorted(reports_dir.iterdir()):
        if not item.is_dir():
            continue
        # Skip COMPLETED and any hidden folders
        if item.name.upper() == "COMPLETED" or item.name.startswith("."):
            continue
        md_files = list(item.glob("*.md"))
        if md_files:
            name = item.name
            if name.endswith("_sections"):
                name = name[: -len("_sections")]
            name = name.replace("_", " ").strip()
            companies.append({
                "name": name,
                "folder": item,
                "section_count": len(md_files),
            })
    return companies


# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê
# 4. SECTION MATCHER ‚Äî fuzzy matches target sections to actual files
# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê

# Common aliases in Indian annual reports
SECTION_ALIASES = {
    "independent audit report": ["independent auditor's report", "auditor's report", "auditors report"],
    "independent auditor report": ["independent auditor's report", "auditor's report"],
    "notes to accounts": ["notes to financial statements", "notes on financial statements", "significant accounting policies"],
    "profit and loss account": ["statement of profit and loss", "profit and loss statement", "income statement"],
    "profit and loss": ["statement of profit and loss"],
    "balance sheet": ["balance sheet"],
    "cash flow": ["cash flow statement"],
    "directors report": ["directors' report", "director's report"],
    "corporate governance": ["corporate governance report"],
    "secretarial audit": ["secretarial audit report"],
    "management discussion": ["management discussion and analysis", "mda"],
}

STOP_WORDS = {"the", "of", "and", "to", "a", "in", "for", "on", "at", "as", "by", "an", "its"}


def _normalize(text: str) -> str:
    return re.sub(r"[^a-z0-9\s]", "", text.lower()).strip()


def match_section(target: str, available: Dict[str, str]) -> List[str]:
    """
    Given a target section name from the rules Excel, find the best-matching
    section(s) from the actual extracted report.

    Returns list of matching section names (may return multiple for broad targets).
    """
    target_norm = _normalize(target)
    available_norm = {_normalize(k): k for k in available}

    # 1. Exact match
    if target_norm in available_norm:
        return [available_norm[target_norm]]

    # 2. Alias expansion
    for alias_key, alias_list in SECTION_ALIASES.items():
        if target_norm in _normalize(alias_key) or _normalize(alias_key) in target_norm:
            for alias in alias_list:
                alias_n = _normalize(alias)
                for avail_n, orig_name in available_norm.items():
                    if alias_n in avail_n or avail_n in alias_n:
                        return [orig_name]

    # 3. Substring match
    matches = []
    for avail_n, orig_name in available_norm.items():
        if target_norm in avail_n or avail_n in target_norm:
            matches.append(orig_name)
    if matches:
        return matches

    # 4. Keyword overlap (at least 2 meaningful words)
    target_words = set(target_norm.split()) - STOP_WORDS
    best = []
    best_score = 0
    for avail_n, orig_name in available_norm.items():
        avail_words = set(avail_n.split()) - STOP_WORDS
        overlap = len(target_words & avail_words)
        if overlap > best_score and overlap >= 2:
            best_score = overlap
            best = [orig_name]
        elif overlap == best_score and overlap >= 2:
            best.append(orig_name)

    return best


def retrieve_sections(available: Dict[str, str], targets: List[str], max_chars: int = 10000) -> tuple:
    """
    Fetch content for the requested target sections.
    Includes clear headers and truncation for token management.
    """
    parts = []
    matched = []
    unmatched = []

    for target in targets:
        found = match_section(target, available)
        if found:
            for section_name in found:
                content = available[section_name]
                if len(content) > max_chars:
                    content = content[:max_chars] + "\n\n[... truncated for context window ...]"
                parts.append(f"‚ïê‚ïê‚ïê {section_name} ‚ïê‚ïê‚ïê\n\n{content}")
                matched.append(section_name)
        else:
            unmatched.append(target)

    result = "\n\n" + "\n\n".join(parts) if parts else ""

    if unmatched:
        result += f"\n\n‚ö† SECTIONS NOT FOUND: {unmatched}"
        result += f"\n  Available sections: {list(available.keys())}"

    return result, matched, unmatched


# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê
# 5. RULES PARSER ‚Äî reads audit rules from Excel
# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê

def parse_question_line(line: str) -> Dict:
    """Parse '1. Question text? : [Section1], [Section2],' ‚Üí structured dict."""
    cleaned = re.sub(r"^\d+\.\s*", "", line.strip())
    targets = re.findall(r"\[([^\]]+)\]", cleaned)
    question = re.sub(r"\s*:?\s*\[.*", "", cleaned).rstrip(",").strip()
    return {"question": question, "target_sections": targets, "financial_check": None}


def parse_financial_check(text: str) -> Dict | None:
    """Parse the Numerator/Denominator financial check block."""
    if "numerator" not in text.lower():
        return None
    num = re.search(r"numerator\s*[-‚Äì:]\s*(.+?)(?=denominator|\Z)", text, re.I | re.DOTALL)
    den = re.search(r"denominator\s*[-‚Äì:]\s*(.+)", text, re.I | re.DOTALL)
    num_secs = re.findall(r"\[([^\]]+)\]", num.group(1)) if num else []
    den_secs = re.findall(r"\[([^\]]+)\]", den.group(1)) if den else []
    return {
        "numerator_desc": num.group(1).strip() if num else "",
        "denominator_desc": den.group(1).strip() if den else "",
        "all_sections": list(set(num_secs + den_secs)),
    }


def parse_rules(excel_path: str, sheet: str = "AS to Validate") -> List[Dict]:
    """Read rules Excel ‚Üí list of structured rule dicts."""
    df = pd.read_excel(excel_path, sheet_name=sheet)
    df.columns = [str(c).strip() for c in df.columns]
    rules = []

    for _, row in df.iterrows():
        rule_id = str(row.get("Rule ID", "")).strip()
        if not rule_id:
            continue

        raw_qs = str(row.get("Audit Questions : Target Sections", ""))
        trigger = str(row.get("Failure Trigger", ""))

        blocks = re.split(r"(?=\d+\.\s)", raw_qs)
        blocks = [b.strip() for b in blocks if b.strip()]

        parsed_qs = []
        fin_check = None
        for block in blocks:
            if block.lower().startswith("the financial check"):
                fin_check = parse_financial_check(block)
            else:
                pq = parse_question_line(block)
                if pq["question"]:
                    parsed_qs.append(pq)

        # Attach financial check to the relevant question
        if fin_check and parsed_qs:
            for pq in reversed(parsed_qs):
                if "financial check" in pq["question"].lower() or "%" in pq["question"]:
                    pq["financial_check"] = fin_check
                    pq["target_sections"] = list(set(pq["target_sections"] + fin_check["all_sections"]))
                    break

        rules.append({
            "rule_id": rule_id,
            "rule_details": str(row.get("Rule Details", "")).strip(),
            "standard_ref": str(row.get("Standard Ref", "")).strip(),
            "audit_requirement": str(row.get("Audit Requirement", "")).strip(),
            "parsed_questions": parsed_qs,
            "failure_trigger_text": trigger,
        })

    return rules


# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê
# 6. LLM UTILITY
# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê

def llm_json(prompt: str, fallback: Dict) -> Dict:
    """Invoke LLM, extract JSON. Returns fallback on any failure."""
    try:
        resp = LLM.invoke([SystemMessage(content=prompt)])
        text = resp.content.strip()
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        return json.loads(text)
    except Exception as e:
        print(f"      ‚ö† LLM/JSON error: {e}")
        return fallback


# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê
# 7. LANGGRAPH NODES
# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê

# ‚îÄ‚îÄ‚îÄ NODE 1: Initialize ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

def node_init(state: AuditState) -> AuditState:
    state["current_question_index"] = 0
    state["question_responses"] = []
    state["failure_trigger_met"] = False
    state["compliance_status"] = None
    state["summary_finding"] = None
    state["auditor_oversight"] = None
    state["confidence_score"] = None
    return state


# ‚îÄ‚îÄ‚îÄ NODE 2: Process Question ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

def node_process_question(state: AuditState) -> AuditState:
    """
    Core reasoning node. For the current question:
      1. Retrieves the right section(s) from the report
      2. Builds context from all prior Q&A (enables cross-section reasoning)
      3. Sends to LLM for analysis
    """
    idx = state["current_question_index"]
    q = state["parsed_questions"][idx]
    total = len(state["parsed_questions"])

    print(f"    Q{idx+1}/{total}: {q['question'][:90]}...")

    # ‚îÄ‚îÄ Retrieve target sections ‚îÄ‚îÄ
    section_content, matched, unmatched = retrieve_sections(
        state["available_sections"], q["target_sections"]
    )

    # ‚îÄ‚îÄ Cumulative context from prior questions (cross-section reasoning) ‚îÄ‚îÄ
    prior_context = ""
    if state["question_responses"]:
        prior_context = "\n\n‚îÄ‚îÄ‚îÄ PRIOR FINDINGS (from earlier steps in this audit) ‚îÄ‚îÄ‚îÄ\n"
        for i, r in enumerate(state["question_responses"]):
            prior_context += (
                f"\nStep {i+1}: {r['question']}\n"
                f"  Answer: {r['answer']}\n"
                f"  Key Evidence: {r['evidence'][:300]}\n"
            )

    # ‚îÄ‚îÄ Financial check instructions (if applicable) ‚îÄ‚îÄ
    fin_block = ""
    if q.get("financial_check"):
        fc = q["financial_check"]
        fin_block = f"""
‚ïê‚ïê‚ïê THIS QUESTION REQUIRES A FINANCIAL CALCULATION ‚ïê‚ïê‚ïê
Numerator: {fc['numerator_desc']}
Denominator: {fc['denominator_desc']}

Procedure:
  1. Extract the relevant finance cost / interest expense figure
  2. Extract borrowings (opening + closing balances)
  3. Average borrowings = (Opening + Closing) / 2
  4. Ratio = (Finance Cost / Average Borrowings) √ó 100
  5. If ratio < 6%, answer YES (suspiciously low)
  6. Show all numbers and the calculation in your reasoning
‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê
"""

    # ‚îÄ‚îÄ Prompt ‚îÄ‚îÄ
    prompt = f"""You are an expert Indian financial auditor analyzing a company's annual report
for compliance with accounting standards and regulatory requirements.

COMPANY: {state['company_name']}
RULE: {state['rule_id']} ‚Äî {state['rule_details']}
STANDARD: {state['standard_ref']}
REQUIREMENT: {state['audit_requirement']}

‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
CURRENT QUESTION ({idx+1} of {total}):
{q['question']}
‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
{prior_context}
{fin_block}
‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
REPORT CONTENT (from sections: {matched}):
{section_content}
‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

INSTRUCTIONS:
- Answer based ONLY on what the report content states or clearly omits
- If information is not found, state "Not found in the provided sections"
- Provide EXACT evidence ‚Äî quote relevant text, table rows, or figures
- Build on prior findings when the current question depends on earlier answers
- Be precise about which section and what part of it you are referencing

Respond ONLY with valid JSON (no markdown fences, no extra text):
{{
    "answer": "Yes / No / Not Found / descriptive answer",
    "evidence": "Exact quote or data from the report supporting your answer",
    "section_ref": "Name of section where evidence was found",
    "reasoning": "Step-by-step explanation of your analysis",
    "confidence": 85
}}"""

    result = llm_json(prompt, {
        "answer": "Error ‚Äî unable to process",
        "evidence": "",
        "section_ref": "",
        "reasoning": "LLM invocation failed",
        "confidence": 0,
    })

    result["question"] = q["question"]
    result["target_sections"] = q["target_sections"]
    result["matched_sections"] = matched
    result["unmatched_sections"] = unmatched
    state["question_responses"].append(result)

    print(f"      ‚Üí {result['answer'][:70]}  [confidence: {result.get('confidence', '?')}]")
    return state


# ‚îÄ‚îÄ‚îÄ NODE 3: Validate ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

def node_validate(state: AuditState) -> AuditState:
    """Senior auditor review ‚Äî checks evidence quality and reasoning logic."""
    idx = state["current_question_index"]
    r = state["question_responses"][idx]

    prompt = f"""You are a SENIOR AUDITOR reviewing a junior analyst's work on an Indian company audit.

Question: {r['question']}
Answer given: {r['answer']}
Evidence cited: {r['evidence']}
Reasoning: {r['reasoning']}
Confidence claimed: {r.get('confidence', 'N/A')}

Evaluate:
1. Is the evidence specific, verifiable, and directly relevant?
2. Does the reasoning logically support the conclusion?
3. Are there gaps or overlooked considerations?
4. Is the confidence score justified given the evidence strength?

Respond ONLY with valid JSON:
{{
    "is_valid": true,
    "issues": "Any problems (or 'None')",
    "adjusted_confidence": {r.get('confidence', 70)}
}}"""

    v = llm_json(prompt, {"is_valid": True, "issues": "Validation skipped", "adjusted_confidence": r.get("confidence", 60)})
    state["question_responses"][idx]["validation"] = v
    if v.get("adjusted_confidence"):
        state["question_responses"][idx]["confidence"] = v["adjusted_confidence"]
    return state


# ‚îÄ‚îÄ‚îÄ NODE 4: Increment ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

def node_increment(state: AuditState) -> AuditState:
    state["current_question_index"] += 1
    return state


# ‚îÄ‚îÄ‚îÄ NODE 5: Evaluate Failure Triggers ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

def node_evaluate_triggers(state: AuditState) -> AuditState:
    """
    After all questions are answered, apply the failure trigger logic
    from the rules Excel to determine final compliance status.
    """
    print(f"  ‚öñ Evaluating failure triggers...")

    qa_block = ""
    for i, r in enumerate(state["question_responses"]):
        qa_block += (
            f"Q{i+1}: {r['question']}\n"
            f"Answer: {r['answer']}\n"
            f"Evidence: {r['evidence']}\n"
            f"Section: {r.get('section_ref', 'N/A')}\n"
            f"Confidence: {r.get('confidence', 'N/A')}\n\n"
        )

    prompt = f"""You are a senior auditor making the FINAL compliance determination for an Indian company.

COMPANY: {state['company_name']}
RULE: {state['rule_id']} ‚Äî {state['rule_details']}
STANDARD: {state['standard_ref']}

‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê
COMPLETE AUDIT TRAIL:
‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê
{qa_block}

‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê
FAILURE TRIGGER LOGIC (apply this EXACTLY):
‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê
{state['failure_trigger_text']}
‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê

INSTRUCTIONS:
- Evaluate each condition in the failure trigger against the evidence gathered
- If ANY failure condition is met ‚Üí "Non-Compliant"
- If a FLAG FOR CHECKING condition is met ‚Üí "Partial" (needs manual review)
- If NO failure conditions are met ‚Üí "Compliant"
- If evidence is insufficient to determine ‚Üí "Partial"
- For Auditor Oversight: did the statutory auditor catch or miss this issue?

Respond ONLY with valid JSON:
{{
    "compliance_status": "Compliant / Non-Compliant / Partial",
    "failure_trigger_met": true or false,
    "summary_finding": "2-3 sentence finding for this rule",
    "auditor_oversight": "Yes ‚Äî [what was missed] / No ‚Äî auditor addressed this adequately",
    "confidence": 85
}}"""

    result = llm_json(prompt, {
        "compliance_status": "Partial",
        "failure_trigger_met": False,
        "summary_finding": "Unable to determine ‚Äî processing error",
        "auditor_oversight": "Unknown",
        "confidence": 50,
    })

    state["failure_trigger_met"] = result.get("failure_trigger_met", False)
    state["compliance_status"] = result.get("compliance_status", "Partial")
    state["summary_finding"] = result.get("summary_finding", "")
    state["auditor_oversight"] = result.get("auditor_oversight", "")
    state["confidence_score"] = result.get("confidence", 50)

    status = state["compliance_status"]
    conf = state["confidence_score"]
    print(f"  ‚Üí {status} (confidence: {conf})")
    return state


# ‚îÄ‚îÄ‚îÄ NODE 6: Assemble Output ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

def node_build_output(state: AuditState) -> AuditState:
    """Assemble all fields for the Compliance Matrix row."""

    # Reasoning path: full step-by-step trail
    rp = []
    for i, r in enumerate(state["question_responses"]):
        rp.append(
            f"Step {i+1}: {r['question']}\n"
            f"  Answer: {r['answer']}\n"
            f"  Reasoning: {r.get('reasoning', 'N/A')}"
        )
    state["reasoning_path"] = "\n\n".join(rp)

    # Evidence: combined from all questions
    ev = []
    for r in state["question_responses"]:
        e = r.get("evidence", "").strip()
        if e and e != "Not found in the provided sections":
            ev.append(e)
    state["evidence_snippet"] = "\n---\n".join(ev) if ev else "No direct evidence found"

    # Section references
    refs = []
    for r in state["question_responses"]:
        ref = r.get("section_ref", "")
        if ref and ref not in refs:
            refs.append(str(ref))
    state["page_ref"] = ", ".join(refs) if refs else "N/A"

    return state


# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê
# 8. GRAPH CONSTRUCTION
# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê

def route_after_increment(state: AuditState) -> str:
    if state["current_question_index"] >= len(state["parsed_questions"]):
        return "evaluate_triggers"
    return "process_question"


def build_graph():
    """
    Compile the LangGraph workflow:

      init ‚îÄ‚îÄ‚ñ∫ process_question ‚îÄ‚îÄ‚ñ∫ validate ‚îÄ‚îÄ‚ñ∫ increment ‚îÄ‚îÄ‚îê
                    ‚ñ≤                                         ‚îÇ
                    ‚îî‚îÄ‚îÄ‚îÄ‚îÄ (more questions?) ‚óÑ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îò
                                ‚îÇ (all done)
                                ‚ñº
                     evaluate_triggers ‚îÄ‚îÄ‚ñ∫ build_output ‚îÄ‚îÄ‚ñ∫ END
    """
    g = StateGraph(AuditState)

    g.add_node("init", node_init)
    g.add_node("process_question", node_process_question)
    g.add_node("validate", node_validate)
    g.add_node("increment", node_increment)
    g.add_node("evaluate_triggers", node_evaluate_triggers)
    g.add_node("build_output", node_build_output)

    g.set_entry_point("init")
    g.add_edge("init", "process_question")
    g.add_edge("process_question", "validate")
    g.add_edge("validate", "increment")
    g.add_conditional_edges("increment", route_after_increment, {
        "process_question": "process_question",
        "evaluate_triggers": "evaluate_triggers",
    })
    g.add_edge("evaluate_triggers", "build_output")
    g.add_edge("build_output", END)

    return g.compile()


# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê
# 9. COMPLIANCE MATRIX WRITER
# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê

COLUMNS = [
    "Company Name", "Rule ID", "Compliance Status", "Summary Finding",
    "Auditor Oversight", "Reasoning Path", "Evidence Snippet",
    "Page Ref", "Confidence Score",
]

HDR_FILL = PatternFill("solid", fgColor="2F5496")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
BODY_FONT = Font(size=10, name="Arial")
WRAP = Alignment(wrap_text=True, vertical="top")
BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
STATUS_FILL = {
    "Compliant": PatternFill("solid", fgColor="C6EFCE"),
    "Non-Compliant": PatternFill("solid", fgColor="FFC7CE"),
    "Partial": PatternFill("solid", fgColor="FFEB9C"),
}


def state_to_row(s: AuditState) -> Dict:
    return {
        "Company Name": s.get("company_name", ""),
        "Rule ID": s.get("rule_id", ""),
        "Compliance Status": s.get("compliance_status", "Partial"),
        "Summary Finding": s.get("summary_finding", ""),
        "Auditor Oversight": s.get("auditor_oversight", ""),
        "Reasoning Path": s.get("reasoning_path", ""),
        "Evidence Snippet": s.get("evidence_snippet", ""),
        "Page Ref": s.get("page_ref", ""),
        "Confidence Score": s.get("confidence_score", 0),
    }


def write_matrix(rows: List[Dict], path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Matrix"
    widths = [25, 14, 18, 50, 35, 60, 50, 22, 15]

    for ci, (col, w) in enumerate(zip(COLUMNS, widths), 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font, c.fill, c.border = HDR_FONT, HDR_FILL, BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[c.column_letter].width = w

    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(COLUMNS, 1):
            val = row.get(col, "")
            c = ws.cell(row=ri, column=ci, value=val)
            c.font, c.alignment, c.border = BODY_FONT, WRAP, BORDER
            if col == "Compliance Status" and str(val).strip() in STATUS_FILL:
                c.fill = STATUS_FILL[str(val).strip()]
                c.font = Font(bold=True, size=10, name="Arial")
            if col == "Confidence Score":
                c.alignment = Alignment(horizontal="center", vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    print(f"\n‚úÖ Compliance Matrix saved ‚Üí {path}")


# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê
# 10. MAIN PIPELINE
# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê

def run(reports_dir: str, rules_path: str, output_path: str, rules_sheet: str = "AS to Validate"):
    reports_root = Path(reports_dir)
    completed_dir = reports_root / "COMPLETED"
    completed_dir.mkdir(parents=True, exist_ok=True)

    companies = discover_companies(reports_root)

    if not companies:
        print(f"No company folders found in {reports_root}")
        print(f"  (Skipping 'COMPLETED' folder)")
        print(f"  Expected: subfolders containing .md section files")
        return

    print(f"Found {len(companies)} company folder(s):")
    for c in companies:
        print(f"  ‚Ä¢ {c['name']} ({c['section_count']} sections) ‚Üí {c['folder'].name}/")

    rules = parse_rules(rules_path, rules_sheet)
    print(f"\nLoaded {len(rules)} rule(s) from '{rules_path}'")
    for r in rules:
        print(f"  ‚Ä¢ {r['rule_id']}: {r['rule_details'][:60]}... ({len(r['parsed_questions'])} questions)")

    workflow = build_graph()
    all_rows = []

    for company in companies:
        sections = load_company_sections(company["folder"])
        print(f"\n{'‚îÅ' * 70}")
        print(f"COMPANY: {company['name']}")
        print(f"  Sections: {list(sections.keys())}")
        print(f"{'‚îÅ' * 70}")

        for rule in rules:
            if not rule["parsed_questions"]:
                print(f"  ‚è≠ {rule['rule_id']}: no questions parsed, skipping")
                continue

            print(f"\n  ‚îÄ‚îÄ Rule {rule['rule_id']}: {rule['rule_details'][:50]}... ‚îÄ‚îÄ")

            initial: AuditState = {
                "company_name": company["name"],
                "rule_id": rule["rule_id"],
                "rule_details": rule["rule_details"],
                "standard_ref": rule["standard_ref"],
                "audit_requirement": rule["audit_requirement"],
                "parsed_questions": rule["parsed_questions"],
                "current_question_index": 0,
                "available_sections": sections,
                "section_filenames": {k: f"{k}.md" for k in sections},
                "question_responses": [],
                "failure_trigger_text": rule["failure_trigger_text"],
                "failure_trigger_met": False,
                "compliance_status": None,
                "summary_finding": None,
                "auditor_oversight": None,
                "reasoning_path": None,
                "evidence_snippet": None,
                "page_ref": None,
                "confidence_score": None,
            }

            final = workflow.invoke(initial)
            all_rows.append(state_to_row(final))

        # ‚îÄ‚îÄ Move completed company folder to COMPLETED ‚îÄ‚îÄ
        dest = completed_dir / company["folder"].name
        if dest.exists():
            shutil.rmtree(dest)  # overwrite if re-running
        shutil.move(str(company["folder"]), str(dest))
        print(f"\n  üìÅ Moved ‚Üí COMPLETED/{company['folder'].name}")

    write_matrix(all_rows, output_path)

    debug = Path(output_path).with_suffix(".debug.json")
    with open(debug, "w", encoding="utf-8") as f:
        json.dump(all_rows, f, indent=2, ensure_ascii=False)
    print(f"  Debug JSON ‚Üí {debug}")

    print(f"\n{'‚ïê' * 70}")
    print(f"DONE ‚Äî {len(companies)} companies √ó {len(rules)} rules = {len(all_rows)} evaluations")
    print(f"{'‚ïê' * 70}")


# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê
# 11. CLI ‚Äî Smart defaults, no required arguments
# ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê

def _auto_detect_paths() -> Dict[str, str]:
    """
    Auto-detect project paths relative to this script's location.
    Expected structure:
      project_root/
        Python/
          audit_orchestration.py   ‚Üê this script
        Extracted_Reports/
          CompanyA_sections/
          COMPLETED/
        Accounting Standard - Rules.xlsx
        Compliance Matrix.xlsx
    """
    script_dir = Path(__file__).resolve().parent        # Python/
    project_root = script_dir.parent                     # project_root/

    # Try multiple common locations for each path
    reports_candidates = [
        project_root / "Extracted_Reports",
        script_dir / "Extracted_Reports",
        Path("Extracted_Reports"),
    ]
    rules_candidates = [
        project_root / "Accounting Standard - Rules.xlsx",
        script_dir / "Accounting Standard - Rules.xlsx",
        Path("Accounting Standard - Rules.xlsx"),
    ]

    reports = next((p for p in reports_candidates if p.exists()), reports_candidates[0])
    rules = next((p for p in rules_candidates if p.exists()), rules_candidates[0])
    output = project_root / "Compliance Matrix.xlsx"

    return {
        "reports": str(reports),
        "rules": str(rules),
        "output": str(output),
    }


if __name__ == "__main__":
    defaults = _auto_detect_paths()

    p = argparse.ArgumentParser(
        description="NFRA Audit Compliance ‚Äî LangGraph Orchestration",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "If no arguments are provided, paths are auto-detected from the project structure.\n"
            f"  --reports  ‚Üí {defaults['reports']}\n"
            f"  --rules    ‚Üí {defaults['rules']}\n"
            f"  --output   ‚Üí {defaults['output']}"
        ),
    )
    p.add_argument("--reports", default=defaults["reports"], help="Extracted_Reports folder")
    p.add_argument("--rules", default=defaults["rules"], help="Accounting Standard Rules Excel")
    p.add_argument("--output", default=defaults["output"], help="Output Compliance Matrix Excel")
    p.add_argument("--sheet", default="AS to Validate", help="Sheet name in rules Excel")
    args = p.parse_args()

    print(f"{'‚ïê' * 70}")
    print(f"NFRA Audit Compliance ‚Äî LangGraph Orchestration Engine")
    print(f"{'‚ïê' * 70}")
    print(f"  Reports:  {args.reports}")
    print(f"  Rules:    {args.rules}")
    print(f"  Output:   {args.output}")
    print(f"  Sheet:    {args.sheet}")
    print(f"{'‚ïê' * 70}\n")

    run(args.reports, args.rules, args.output, args.sheet)